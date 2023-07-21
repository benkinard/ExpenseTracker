"""Define classes for connecting and writing to an Income/Expense Tracker Excel Workbook

Classes
-------
TrackerSection
    Defines a group of cells in the Income/Excel Tracker Excel Workbook that are used for a particular category
IncomeExpenseTracker
    Represents a page from the Income & Expense Tracker Excel Workbook corresponding to a specific month & year
"""
import logging
import openpyxl
import pandas as pd
import tracker.exceptions
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from tracker.resources import PROJECT_ROOT_PATH
from transaction.transactions import Transactions
from transaction.dao.transaction_dao import FlatFileTransactionDAO
from typing import Callable, Optional


class TrackerSection:
    """
    Class representing a collection of cells within the Income & Expense Tracker Excel Workbook to catalog a category
    of transactions

    Instance Variables
    ------------------
    name : str
        Category name
    xl_worksheet : Worksheet
        Excel worksheet the collection of cells belong to
    trx_type : str, "income" or "expense"
        Designates the transactions that make up this section as either income or expense
    keywords : list[str]
        List of keywords to search for in transaction descriptions
    min_row : int
        Row number the section starts at
    max_row : int
        Row number the section ends at
    min_col : int
        Column number the section starts at
    max_col : int
        Column number the section ends at
    keyword_exceptions : list[str]
        List of keywords that should disqualify a transaction from belonging to the category
    is_inverse_section : bool
        Whether the section should treat keywords as qualifiers for section membership or dis-qualifiers

    Public Methods
    --------------
    clear_contents(self)
        Delete data from every cell in the section
    write_data(self, transactions: pd.DataFrame)
        Write transaction data to the cells in the section
    """

    def __init__(self, name: str, xl_worksheet: Worksheet, trx_type: str, keywords: list[str], min_row: int,
                 max_row: int, min_col: int, max_col: int, keyword_exceptions: list[str] = None,
                 is_inverse_section: bool = False):
        """Constructor for TrackerSection Class

        Parameters
        ----------
        name : str
            Category name
        xl_worksheet : Worksheet
            Excel worksheet the collection of cells belong to
        trx_type : str, "income" or "expense"
            Designates the transactions that make up this section as either income or expense
        keywords : list[str]
            List of keywords to search for in transaction descriptions
        min_row : int
            Row number the section starts at
        max_row : int
            Row number the section ends at
        min_col : int
            Column number the section starts at
        max_col : int
            Column number the section ends at
        keyword_exceptions : list[str]
            List of keywords that should disqualify a transaction from belonging to the category
        is_inverse_section : bool
            Whether the section should treat keywords as qualifiers for section membership or dis-qualifiers

        Raises
        ------
        tracker.exceptions.InvalidTrxType
            If trx_type is not one of either "expense" or "income"
        """
        if trx_type not in ["expense", "income"]:
            raise tracker.exceptions.InvalidTrxType(f"TrackerSection requires \"expense\" or \"income\" for trx_type, "
                                                    f"got \"{trx_type}\"")
        # Public instance variables
        self.name: str = name
        self.xl_worksheet: Worksheet = xl_worksheet
        self.trx_type: str = trx_type
        self.keywords: list[str] = keywords
        self.min_row: int = min_row
        self.max_row: int = max_row
        self.min_col: int = min_col
        self.max_col: int = max_col
        self.keyword_exceptions: list[str] = [] if keyword_exceptions is None else keyword_exceptions
        self.is_inverse_section: bool = is_inverse_section

    # Public methods
    def clear_contents(self):
        """Deletes data from every cell in the TrackerSection"""
        for row in self.xl_worksheet.iter_rows(min_row=self.min_row, min_col=self.min_col, max_row=self.max_row,
                                               max_col=self.max_col):
            for cell in row:
                if isinstance(cell, Cell):
                    cell.value = None

    def write_data(self, transactions: pd.DataFrame):
        """Write transaction data to the cells of the TrackerSection

        Parameters
        ----------
        transactions : pd.DataFrame
            Table of either all income or all expense transactions for the period

        Raises
        ------
        tracker.exceptions.InsufficientTrackerSectionSize
            If the number of transactions that belong to the TrackerSection exceed the number of rows allotted to the
            TrackerSection
        """
        # From transaction data passed into the method, filter for transactions relevant to this section only
        section_trx: pd.DataFrame = transactions.loc[list(map(self.__trx_filter(), transactions['Description'])), :]
        section_trx.reset_index(drop=True, inplace=True)
        # Confirm the section is large enough to display the transactions that belong to it
        if len(section_trx) > (self.max_row - self.min_row + 1):
            raise tracker.exceptions.InsufficientTrackerSectionSize(f"TrackerSection \"{self.name}\" transactions "
                                                                    f"({len(section_trx)}) exceeds row allowance "
                                                                    f"({self.max_row - self.min_row + 1})")

        # The Date, Description, and Amount columns begin at the section's first, second, and last columns respectively
        date_column = chr(64 + self.min_col)
        desc_column = chr(64 + self.min_col + 1)
        amt_column = chr(64 + self.max_col)
        # For each transaction belonging to this section, write info to their respective columns
        for idx, row in section_trx.iterrows():
            self.xl_worksheet[f"{date_column}{str(self.min_row + idx)}"] = row['Posting Date']
            self.xl_worksheet[f"{desc_column}{str(self.min_row + idx)}"] = row['Description']
            self.xl_worksheet[f"{amt_column}{str(self.min_row + idx)}"] = row['Amount']

    # Private methods
    def __trx_filter(self) -> Callable[[str], bool]:
        """
        Defines how to filter the table of all transactions into just the transactions relevant to the
        TrackerSection

        Returns
        -------
        Callable[[str], bool]
            Function that will determine if keywords are found or not found in a string passed in as the argument
        """
        # If this section IS marked as inverse, then filter for transactions that do NOT contain the provided keywords
        if self.is_inverse_section:
            if len(self.keyword_exceptions) == 0:
                return lambda text: all(key_word not in text.upper() for key_word in self.keywords)
            # If there are keyword exceptions, make sure they are not included when filtering
            else:
                return lambda text: all(key_word not in text.upper() for key_word in self.keywords) or \
                                    any(kw_ex in text.upper() for kw_ex in self.keyword_exceptions)
        # If this section is NOT marked as inverse, then filter for transactions that DO contain the provided keywords
        else:
            if len(self.keyword_exceptions) == 0:
                return lambda text: any(key_word in text.upper() for key_word in self.keywords)
            # If there are keyword exceptions, make sure they are not included when filtering
            else:
                return lambda text: (any(key_word in text.upper() for key_word in self.keywords)) and \
                                    all(kw_ex not in text.upper() for kw_ex in self.keyword_exceptions)

    # Magic methods
    def __eq__(self, other) -> bool:
        return self.name == other.name and self.trx_type == other.trx_type \
               and self.min_row == other.min_row and self.max_row == other.max_row \
               and self.min_col == other.min_col and self.max_col == other.max_col \
               and self.keywords == other.keywords and self.keyword_exceptions == other.keyword_exceptions \
               and self.is_inverse_section == other.is_inverse_section


class IncomeExpenseTracker:
    """Class representing a page from the Income & Expense Tracker Excel Workbook for a specific month & year

    Class Variables
    ---------------
    AS_OF_DATE_CELL
        Cell containing the date that the transaction data was last updated
    TRACKER_ROOT_PATH
        File path to the directory containing the Income & Expense Tracker Excel Workbooks and raw transaction data

    Instance Variables
    ------------------
    month_name : str
        Name of the month the instance corresponds to
    month_num : int
        Number (1-12) of the month the instance corresponds to
    year : int
        Year the instance corresponds to

    Public Methods
    --------------
    add_section(self, name, keywords, min_row, max_row, min_col, max_col, trx_type, keyword_exceptions,
    is_inverse_section)
        Add instance of TrackerSection to the Income & Expense Tracker
    close_tracker(self)
        Close the connection to the Excel Workbook
    delete_section(self, name)
        Remove an instance of TrackerSection from the Income & Expense Tracker
    replace_section(self, name, keywords, min_row, max_row, min_col, max_col, trx_type, keyword_exceptions,
    is_inverse_section)
        Replace a specified instance of TrackerSection in the Income & Expense Tracker with another one
    update_tracker(self)
        Update transaction data in all sections of the Income & Expense Tracker
    """
    # Class variables
    AS_OF_DATE_CELL = "T1"
    TRACKER_ROOT_PATH = f"{PROJECT_ROOT_PATH}/Tracker"

    def __init__(self, month_name: str, month_num: int, year: int):
        """Constructor for IncomeExpenseTracker Class

        Parameters
        ----------
        month_name : str
            Name of the month the instance corresponds to
        month_num : int
            Number (1-12) of the month the instance corresponds to
        year : int
            Year the instance corresponds to

        Raises
        ------
        FileNotFoundError
            If an Income & Expense Tracker Excel Workbook with the provided year cannot be found
        KeyError
            If a sheet from the Income & Expense Tracker Excel Workbook cannot be found for the provided month
        """
        # Public instance variables
        self.month_name: str = month_name
        self.month_num: int = month_num
        self.year: int = year
        # Private instance variables
        self.__transactions: Optional[Transactions] = None
        self.__xl_workbook_path: str = f"{self.__class__.TRACKER_ROOT_PATH}/{year}/Income&Expenses{year}.xlsx"
        try:
            self.__xl_workbook: Workbook = openpyxl.load_workbook(self.__xl_workbook_path)
            self.__xl_worksheet: Worksheet = self.__xl_workbook[f"{month_name} {year}"]
        except FileNotFoundError as fnf_err:
            raise tracker.exceptions.TrackerWorkbookDoesNotExist(fnf_err)
        except KeyError as ke:
            raise tracker.exceptions.TrackerWorksheetDoesNotExist(ke)
        self.__sections: dict[str, TrackerSection] = {}

    # Public methods
    def add_section(self, name: str, keywords: list[str], min_row: int, max_row: int, min_col: int, max_col: int,
                    trx_type: str = "expense", keyword_exceptions: list[str] = None, is_inverse_section: bool = False):
        """Add a section for recording a group of related transactions in the tracker

        Parameters
        ----------
        name : str
            Category name
        keywords : list[str]
            List of keywords to search for in transaction descriptions
        min_row : int
            Row number the section starts at
        max_row : int
            Row number the section ends at
        min_col : int
            Column number the section starts at
        max_col : int
            Column number the section ends at
        trx_type : str, default "expense"
            Designates the transactions that make up this section as either income or expense
        keyword_exceptions : list[str], default None
            List of keywords that should disqualify a transaction from belonging to the category
        is_inverse_section : bool, default False
            Whether the section should treat keywords as qualifiers for section membership or dis-qualifiers

        Raises
        ------
        tracker.exceptions.TrackerSectionAlreadyExists
            If the section attempting to be added already exists in the tracker
        """
        if name in self.__sections.keys():
            raise tracker.exceptions.TrackerSectionAlreadyExists(f"Section \"{name}\" already exists. If you'd like to "
                                                                 f"replace the existing section, make a call to the "
                                                                 f"\"replace_section()\" method instead.")

        section = TrackerSection(name, self.__xl_worksheet, trx_type, keywords, min_row, max_row, min_col, max_col,
                                 keyword_exceptions, is_inverse_section)
        self.__sections[section.name] = section

    def close_tracker(self):
        """Close connection to the Income & Expense Tracker Excel Workbook"""
        if self.__xl_workbook:
            self.__xl_workbook.close()

    def delete_section(self, name: str):
        """Remove TrackerSection with provided name from the tracker

        Parameters
        ----------
        name : str
            Name of the section to be removed

        Raises
        ------
        tracker.exceptions.TrackerSectionDoesNotExist
            If the name provided does not correspond to an existing section within the tracker
        """
        if name not in self.__sections.keys():
            raise tracker.exceptions.TrackerSectionDoesNotExist(f"Cannot delete section \"{name}\" because it does "
                                                                f"not exist")
        del self.__sections[name]

    def replace_section(self, name: str, keywords: list[str], min_row: int, max_row: int, min_col: int, max_col: int,
                        trx_type: str = "expense", keyword_exceptions: list[str] = None,
                        is_inverse_section: bool = False):
        """Replace an existing section within the tracker with a new one

        Parameters
        ----------
        name : str
            Category name
        keywords : list[str]
            List of keywords to search for in transaction descriptions
        min_row : int
            Row number the section starts at
        max_row : int
            Row number the section ends at
        min_col : int
            Column number the section starts at
        max_col : int
            Column number the section ends at
        trx_type : str, default "expense"
            Designates the transactions that make up this section as either income or expense
        keyword_exceptions : list[str], default None
            List of keywords that should disqualify a transaction from belonging to the category
        is_inverse_section : bool, default False
            Whether the section should treat keywords as qualifiers for section membership or dis-qualifiers

        Raises
        ------
        tracker.exceptions.TrackerSectionDoesNotExist
            If the section name does not correspond to an existing section within the tracker
        tracker.exceptions.ReplaceTrackerSectionError
            If the replacement section is identical to the already existing section within the tracker
        """
        if name not in self.__sections.keys():
            raise tracker.exceptions.TrackerSectionDoesNotExist(f"Cannot replace section \"{name}\" because it does "
                                                                f"not exist")

        section = TrackerSection(name, self.__xl_worksheet, trx_type, keywords, min_row, max_row, min_col,
                                 max_col, keyword_exceptions, is_inverse_section)
        if section == self.__sections[name]:
            raise tracker.exceptions.ReplaceTrackerSectionError(f"Replacement TrackerSection is identical to current "
                                                                f"one")
        self.__sections[name] = section

    def update_tracker(self):
        """Update the tracker with the latest transaction data

        Raises
        ------
        tracker.exceptions.EmptyTrackerError
            If the tracker has no sections assigned to it
        """
        if len(self.__sections) == 0:
            raise tracker.exceptions.EmptyTrackerError(f"{self.month_name} {self.year} Income & Expense Tracker "
                                                       f"contains no TrackerSections. Cannot update.")

        logging.info(f"Clearing {self.month_name} {self.year} Income & Expense Tracker Contents")
        self.__clear_tracker_contents()
        logging.info(f"Pulling Updated Transactions Data for {self.month_name} {self.year}")
        self.__get_transactions()
        logging.info(f"Writing Updated {self.month_name} {self.year} Transactions Data to Tracker")
        self.__write_transaction_data_to_tracker()
        logging.info(f"Saving Updates to {self.month_name} {self.year} Income & Expense Tracker")
        self.__xl_workbook.save(self.__xl_workbook_path)

    # Private methods
    def __clear_tracker_contents(self):
        """Delete data from every section within the tracker"""
        self.__xl_worksheet[self.__class__.AS_OF_DATE_CELL] = None
        for section in self.__sections.values():
            section.clear_contents()

    def __get_transactions(self):
        """Pull the transactions for the period"""
        if not self.__transactions:
            self.__transactions = Transactions(self.month_num, self.year,
                                               FlatFileTransactionDAO(self.__class__.TRACKER_ROOT_PATH))

        self.__transactions.get_transactions_for_the_period()

    def __write_transaction_data_to_tracker(self):
        """Write latest transaction data to each section in the tracker"""
        self.__xl_worksheet[self.__class__.AS_OF_DATE_CELL] = self.__transactions.get_as_of_date()
        for section in self.__sections.values():
            # Determine if base transactions the section should filter from should be expense or income transactions
            if section.trx_type == "expense":
                trx = self.__transactions.get_expenses()
            else:
                trx = self.__transactions.get_income()
            section.write_data(trx)
